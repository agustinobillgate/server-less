"""
VHPMobile API Testing Tool
Moves files, maps endpoints, tests APIs, and generates Excel report
"""

import os
import shutil
import json
import time
import requests
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import traceback


class APITester:
    def __init__(self, base_url="http://localhost:8000"):
        self.base_url = base_url
        self.test_payload = {
            "inputUserkey": "95EE44CBF839764A7690C157AC66C9C902905E01",
            "inputUsername": "it",
            "hotel_schema": "qcserverless3"
        }
        self.results = []
        self.moved_files = set()
        self.mapped_functions = {}
        self.unmapped_files = set()
        
    def log(self, message):
        """Print timestamped log message"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] {message}")
    
    def move_files(self, source_dir="converted2/VHPMobile", target_dir="functions"):
        """Move all files from source to target directory (flattened)"""
        self.log("=" * 80)
        self.log("STEP 1: Moving files from converted2/VHPMobile to functions")
        self.log("=" * 80)
        
        source_path = Path(source_dir)
        target_path = Path(target_dir)
        
        # Create target directory if it doesn't exist
        target_path.mkdir(parents=True, exist_ok=True)
        
        if not source_path.exists():
            self.log(f"ERROR: Source directory '{source_dir}' does not exist!")
            return False
        
        moved_count = 0
        error_count = 0
        
        # Walk through all files in source directory
        for root, dirs, files in os.walk(source_path):
            for filename in files:
                source_file = Path(root) / filename
                target_file = target_path / filename
                
                try:
                    # Check if file already exists in target
                    if target_file.exists():
                        self.log(f"  SKIP: {filename} (already exists in target)")
                        self.moved_files.add(filename)
                    else:
                        shutil.move(str(source_file), str(target_file))
                        self.log(f"  MOVED: {filename}")
                        moved_count += 1
                        self.moved_files.add(filename)
                except Exception as e:
                    self.log(f"  ERROR moving {filename}: {str(e)}")
                    error_count += 1
        
        self.log(f"\nFiles moved: {moved_count}")
        self.log(f"Files already in target: {len(self.moved_files) - moved_count}")
        self.log(f"Errors: {error_count}")
        self.log(f"Total files in functions folder: {len(self.moved_files)}")
        return True
    
    def scan_mappings(self, modules_dir="modules/VHPMobile"):
        """Scan all _mapping.txt files in modules directory"""
        self.log("\n" + "=" * 80)
        self.log("STEP 2: Scanning _mapping.txt files")
        self.log("=" * 80)
        
        modules_path = Path(modules_dir)
        
        if not modules_path.exists():
            self.log(f"ERROR: Modules directory '{modules_dir}' does not exist!")
            return False
        
        mapping_count = 0
        
        # Find all subdirectories
        for subdir in modules_path.iterdir():
            if subdir.is_dir():
                mapping_file = subdir / "_mapping.txt"
                
                if mapping_file.exists():
                    self.log(f"\nProcessing: {subdir.name}/_mapping.txt")
                    
                    try:
                        with open(mapping_file, 'r', encoding='utf-8') as f:
                            lines = f.readlines()
                        
                        # Skip header line
                        for line in lines[1:]:
                            line = line.strip()
                            if not line:
                                continue
                            
                            parts = line.split(',')
                            if len(parts) >= 2:
                                service = parts[0].strip()
                                function = parts[1].strip()
                                
                                # Store mapping with module context
                                endpoint = f"/VHPMobile/{subdir.name}/{service}"
                                function_file = f"{function}.py"
                                
                                if endpoint not in self.mapped_functions:
                                    self.mapped_functions[endpoint] = {
                                        'service': service,
                                        'function': function,
                                        'function_file': function_file,
                                        'module': subdir.name
                                    }
                                    mapping_count += 1
                                    self.log(f"  Mapped: {service} -> {function}.py")
                    
                    except Exception as e:
                        self.log(f"  ERROR reading {mapping_file}: {str(e)}")
                        self.log(f"  Traceback: {traceback.format_exc()}")
        
        self.log(f"\nTotal mappings found: {mapping_count}")
        return True
    
    def check_unmapped_files(self):
        """Check for files in functions that aren't mapped"""
        self.log("\n" + "=" * 80)
        self.log("STEP 3: Checking for unmapped files")
        self.log("=" * 80)
        
        # Get all function files from mappings
        mapped_files = {m['function_file'] for m in self.mapped_functions.values()}
        
        # Find unmapped files
        self.unmapped_files = self.moved_files - mapped_files
        
        if self.unmapped_files:
            self.log(f"\nFound {len(self.unmapped_files)} unmapped files:")
            for filename in sorted(self.unmapped_files):
                self.log(f"  - {filename}")
                
                # Add to results as error
                self.results.append({
                    'endpoint': 'N/A',
                    'service': 'N/A',
                    'function_file': filename,
                    'request': json.dumps(self.test_payload, indent=2),
                    'response': 'N/A',
                    'response_time': 'N/A',
                    'error': 'No Mapping - File exists but not listed in any _mapping.txt',
                    'status': 'ERROR'
                })
        else:
            self.log("\nAll files are mapped correctly!")
    
    def test_endpoint(self, endpoint, mapping_info):
        """Test a single endpoint and return result"""
        url = f"{self.base_url}{endpoint}"
        function_file = mapping_info['function_file']
        
        self.log(f"\nTesting: {endpoint}")
        self.log(f"  Function: {function_file}")
        
        result = {
            'endpoint': endpoint,
            'service': mapping_info['service'],
            'function_file': function_file,
            'request': json.dumps(self.test_payload, indent=2),
            'response': '',
            'response_time': '',
            'error': '',
            'status': 'UNKNOWN'
        }
        
        try:
            # Check if function file exists
            function_path = Path("functions") / function_file
            if not function_path.exists():
                result['response'] = json.dumps({
                    "error": f"{function_file} not exists."
                }, indent=2)
                result['error'] = f"Function file '{function_file}' not found in functions folder"
                result['status'] = 'ERROR'
                result['response_time'] = 'N/A'
                self.log(f"  ERROR: Function file not found")
                return result
            
            # Make API request
            start_time = time.time()
            
            try:
                response = requests.post(
                    url,
                    json=self.test_payload,
                    timeout=30,
                    headers={'Content-Type': 'application/json'}
                )
                
                end_time = time.time()
                response_time = round((end_time - start_time) * 1000, 2)  # in milliseconds
                
                result['response_time'] = f"{response_time} ms"
                
                # Parse response
                try:
                    response_json = response.json()
                    result['response'] = json.dumps(response_json, indent=2)
                    
                    # Check success criteria
                    if 'response' in response_json:
                        output_ok = response_json['response'].get('outputOkFlag', '').lower() == 'true'
                        server_error = response_json.get('serverinfo', {}).get('error', '')
                        
                        if output_ok and not server_error:
                            result['status'] = 'SUCCESS'
                            self.log(f"  SUCCESS ({response_time} ms)")
                        else:
                            result['status'] = 'ERROR'
                            if not output_ok:
                                result['error'] = 'outputOkFlag is not true'
                            if server_error:
                                result['error'] = result['error'] + '; ' if result['error'] else ''
                                result['error'] += f"Server error: {server_error}"
                            self.log(f"  ERROR: {result['error']}")
                    
                    # Check for direct error response
                    elif 'error' in response_json:
                        result['status'] = 'ERROR'
                        result['error'] = response_json['error']
                        self.log(f"  ERROR: {result['error']}")
                    
                    else:
                        result['status'] = 'ERROR'
                        result['error'] = 'Response format not recognized'
                        self.log(f"  ERROR: Unexpected response format")
                
                except json.JSONDecodeError as e:
                    result['response'] = response.text[:1000]  # First 1000 chars
                    result['status'] = 'ERROR'
                    result['error'] = f'Invalid JSON response: {str(e)}'
                    self.log(f"  ERROR: Invalid JSON response")
            
            except requests.Timeout:
                result['response'] = 'Request timeout (30s)'
                result['error'] = 'Request timeout after 30 seconds'
                result['status'] = 'ERROR'
                result['response_time'] = 'TIMEOUT'
                self.log(f"  ERROR: Request timeout")
            
            except requests.ConnectionError as e:
                result['response'] = f'Connection error: {str(e)}'
                result['error'] = f'Cannot connect to server: {str(e)}'
                result['status'] = 'ERROR'
                result['response_time'] = 'N/A'
                self.log(f"  ERROR: Connection error")
            
            except Exception as e:
                result['response'] = f'Request error: {str(e)}'
                result['error'] = f'Request failed: {str(e)}'
                result['status'] = 'ERROR'
                result['response_time'] = 'N/A'
                self.log(f"  ERROR: {str(e)}")
        
        except Exception as e:
            result['error'] = f'Unexpected error: {str(e)}'
            result['status'] = 'ERROR'
            self.log(f"  FATAL ERROR: {str(e)}")
            self.log(f"  Traceback: {traceback.format_exc()}")
        
        return result
    
    def test_all_endpoints(self):
        """Test all mapped endpoints"""
        self.log("\n" + "=" * 80)
        self.log("STEP 4: Testing all endpoints")
        self.log("=" * 80)
        
        total = len(self.mapped_functions)
        self.log(f"\nTotal endpoints to test: {total}")
        
        for idx, (endpoint, mapping_info) in enumerate(self.mapped_functions.items(), 1):
            self.log(f"\n[{idx}/{total}] " + "-" * 70)
            result = self.test_endpoint(endpoint, mapping_info)
            self.results.append(result)
            
            # Small delay between requests to avoid overwhelming server
            time.sleep(0.1)
    
    def generate_excel_report(self, filename="api_test_report.xlsx"):
        """Generate Excel report with results"""
        self.log("\n" + "=" * 80)
        self.log("STEP 5: Generating Excel report")
        self.log("=" * 80)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "API Test Results"
            
            # Define headers
            headers = ['Endpoint', 'Service', 'Function File', 'Request', 'Response', 'Response Time', 'Error', 'Status']
            
            # Style for headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data
            for row_num, result in enumerate(self.results, 2):
                ws.cell(row=row_num, column=1, value=result['endpoint'])
                ws.cell(row=row_num, column=2, value=result['service'])
                ws.cell(row=row_num, column=3, value=result['function_file'])
                ws.cell(row=row_num, column=4, value=result['request'])
                ws.cell(row=row_num, column=5, value=result['response'])
                ws.cell(row=row_num, column=6, value=result['response_time'])
                ws.cell(row=row_num, column=7, value=result['error'])
                ws.cell(row=row_num, column=8, value=result['status'])
                
                # Color code status
                status_cell = ws.cell(row=row_num, column=8)
                if result['status'] == 'SUCCESS':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                elif result['status'] == 'ERROR':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")
                
                # Wrap text for JSON columns
                for col in [4, 5, 7]:
                    ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Adjust column widths
            column_widths = {
                'A': 40,  # Endpoint
                'B': 25,  # Service
                'C': 30,  # Function File
                'D': 50,  # Request
                'E': 60,  # Response
                'F': 15,  # Response Time
                'G': 50,  # Error
                'H': 12   # Status
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Set row heights
            ws.row_dimensions[1].height = 30
            for row_num in range(2, len(self.results) + 2):
                ws.row_dimensions[row_num].height = 100
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(filename)
            
            self.log(f"\nReport saved: {filename}")
            
            # Print summary
            success_count = sum(1 for r in self.results if r['status'] == 'SUCCESS')
            error_count = sum(1 for r in self.results if r['status'] == 'ERROR')
            
            self.log("\n" + "=" * 80)
            self.log("SUMMARY")
            self.log("=" * 80)
            self.log(f"Total endpoints tested: {len(self.results)}")
            self.log(f"Success: {success_count}")
            self.log(f"Errors: {error_count}")
            self.log(f"Success rate: {(success_count/len(self.results)*100):.1f}%")
            
            return True
            
        except Exception as e:
            self.log(f"ERROR generating report: {str(e)}")
            self.log(f"Traceback: {traceback.format_exc()}")
            return False
    
    def run(self):
        """Run the complete testing workflow"""
        start_time = datetime.now()
        self.log("=" * 80)
        self.log("VHPMobile API Testing Tool Started")
        self.log(f"Start time: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        self.log("=" * 80)
        
        try:
            # Step 1: Move files
            if not self.move_files():
                self.log("\nERROR: File moving failed. Aborting.")
                return
            
            # Step 2: Scan mappings
            if not self.scan_mappings():
                self.log("\nERROR: Mapping scan failed. Aborting.")
                return
            
            # Step 3: Check unmapped files
            self.check_unmapped_files()
            
            # Step 4: Test endpoints
            self.test_all_endpoints()
            
            # Step 5: Generate report
            self.generate_excel_report()
            
            end_time = datetime.now()
            duration = end_time - start_time
            
            self.log("\n" + "=" * 80)
            self.log("VHPMobile API Testing Tool Completed")
            self.log(f"End time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
            self.log(f"Total duration: {duration}")
            self.log("=" * 80)
            
        except Exception as e:
            self.log(f"\nFATAL ERROR: {str(e)}")
            self.log(f"Traceback: {traceback.format_exc()}")


if __name__ == "__main__":
    tester = APITester(base_url="http://localhost:8000")
    tester.run()